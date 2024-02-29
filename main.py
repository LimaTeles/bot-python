from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

caminho_r1 = '<CAMINHO_DO-ARQUIVO>'
caminho_r2 = '<CAMINHO_DO-ARQUIVO>'
caminho_r3 = '<CAMINHO_DO-ARQUIVO>'
caminho_r4 = '<CAMINHO_DO-ARQUIVO>'
caminho_r5 = '<CAMINHO_DO-ARQUIVO>'

caminho_meta = '<CAMINHO_DO-ARQUIVO>'
caminho_controle = '<CAMINHO_DO-ARQUIVO>'

campo_r1 = load_workbook(caminho_r1)
campo_r2 = load_workbook(caminho_r2)
campo_r3 = load_workbook(caminho_r3)
campo_r4 = load_workbook(caminho_r4)
campo_r5 = load_workbook(caminho_r5)

meta = load_workbook(caminho_meta)
controle = load_workbook(caminho_controle)

class Atualizacao: 
    
    def __init__(self, sheet_municipio, regiao, nome_municipio):
        
        self.sheet_municipio = sheet_municipio
        self.regiao = regiao
        self.nome_municipio = nome_municipio
        
        self.sheet_meta = meta['SHEET_DA-PLANILHA']
        self.sheet_resumo = meta['SHEET_DA-PLANILHA']
        self.sheet_micro = controle['SHEET_DA-PLANILHA']
        self.sheet_controle = controle['SHEET_DA-PLANILHA']
        self.sheet_monitoramento = controle['SHEET_DA-PLANILHA']
        
    def atualizar_meta(self, coluna_meta):
        
        dicionario_meta = {}
        
        for linha in range(2, self.sheet_municipio.max_row + 1):
            if all(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']):
                
                data = self.sheet_municipio['L' + str(linha)].value
                equipe = self.sheet_municipio['M' + str(linha)].value
                
                chave_dicionario_meta = (data, equipe)
                
                if chave_dicionario_meta in dicionario_meta:
                    dicionario_meta[chave_dicionario_meta] += 1
                else:
                    dicionario_meta[chave_dicionario_meta] = 1
        
        resultados = []
        
        for (data, equipe), contagem_coletas in dicionario_meta.items():
            resultados.append((data, contagem_coletas, equipe))
        
        proxima_linha = 1
        
        while self.sheet_meta.cell(row = proxima_linha, column = coluna_meta).value is not None:
            proxima_linha += 1   
            
        for linha, resultado in enumerate(resultados, start = proxima_linha):
            for coluna, valor in enumerate(resultado, start = coluna_meta):
                self.sheet_meta.cell(row = linha, column = coluna, value = valor)
    
    def atualizar_controle(self):
        
        dicionario_controle = {}
        
        for linha in range(2, self.sheet_municipio.max_row + 1):
            if all(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']):
                
                data = self.sheet_municipio['L' + str(linha)].value
                equipe = self.sheet_municipio['M' + str(linha)].value
                comunidades = self.sheet_municipio['I' + str(linha)].value
                
                chave_dicionario_controle = (data, equipe, comunidades)
                
                if chave_dicionario_controle in dicionario_controle:
                    dicionario_controle[chave_dicionario_controle] += 1
                else:
                    dicionario_controle[chave_dicionario_controle] = 1
   
        resultados = []
    
        for (data, equipe, comunidades), contagem_coletas in dicionario_controle.items():
            resultados.append((data, self.regiao, equipe, self.nome_municipio, comunidades, contagem_coletas))
        
        proxima_linha = 1
        
        while self.sheet_controle.cell(row = proxima_linha, column = 2).value is not None:
            proxima_linha += 1
            
        for linha, resultado in enumerate(resultados, start = proxima_linha):
            for coluna, valor in enumerate(resultado, start = 1):
                self.sheet_controle.cell(row = linha, column = coluna, value = valor)  
                
    def atualizar_monitoramento(self):
        
        dicionario_completas = {}
        dicionario_incompletas = {}

        for linha in range(2, self.sheet_municipio.max_row + 1):
            if all(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']):

                data = self.sheet_municipio['L' + str(linha)].value
                equipe = self.sheet_municipio['M' + str(linha)].value
                
                chave_dicionario_completas = (data, equipe)

                if chave_dicionario_completas in dicionario_completas:
                    dicionario_completas[chave_dicionario_completas] += 1
                else:
                    dicionario_completas[chave_dicionario_completas] = 1

            if any(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']) and any(self.sheet_municipio[coluna + str(linha)].value is None for coluna in ['L', 'M', 'N', 'O', 'P']):

                data = self.sheet_municipio['L' + str(linha)].value
                equipe = self.sheet_municipio['M' + str(linha)].value
                
                chave_dicionario_incompletas = (data, equipe)

                if chave_dicionario_incompletas in dicionario_incompletas:
                    dicionario_incompletas[chave_dicionario_incompletas] += 1
                else:
                    dicionario_incompletas[chave_dicionario_incompletas] = 1
          
        resultados = []
        
        for (data, equipe), contagem_completas in dicionario_completas.items():
            resultados.append((data, self.regiao, self.nome_municipio, equipe, contagem_completas, 'Sim'))

        for (data, equipe), contagem_incompletas in dicionario_incompletas.items():
            resultados.append((data, self.regiao, self.nome_municipio, equipe, contagem_incompletas, 'Não')) 

        proxima_linha = 1
        
        while self.sheet_monitoramento.cell(row = proxima_linha, column = 2).value is not None:
            proxima_linha += 1
            
        for linha, resultado in enumerate(resultados, start = proxima_linha):
            for coluna, valor in enumerate(resultado, start = 1):
                self.sheet_monitoramento.cell(row = linha, column = coluna, value = valor)

    def coloracao_incompletas(self):
        
        for linha in range(2, self.sheet_municipio.max_row + 1):
            if any(self.sheet_municipio[coluna + str(linha)].value is not None for coluna in ['L', 'M', 'N', 'O', 'P']) and any(self.sheet_municipio[coluna + str(linha)].value is None for coluna in ['L', 'M', 'N', 'O', 'P']):
               
                red_fill = PatternFill(start_color = 'FF6666', fill_type = 'solid')
                
                for coluna in range(12, 17):
                    
                    self.sheet_municipio.cell(row = linha, column = coluna).fill = red_fill 
                    
    def situacoes(self, linha_resumo):
        
        endereco_nao_localizado = 0
        endereco_outro_municipio = 0
        contato_sem_sucesso = 0
        
        for linha in range(2, self.sheet_municipio.max_row + 1):
            
            situacao = self.sheet_municipio['S' + str(linha)].value
            
            if situacao == 'Endereço não localizado':
                endereco_nao_localizado +=1
            elif situacao == 'Endereço em outro município':
                endereco_outro_municipio += 1
            elif situacao == 'Contato sem sucesso':
                contato_sem_sucesso += 1
        
        resultados = [endereco_outro_municipio, endereco_nao_localizado, contato_sem_sucesso]
            
        for coluna, valor in zip('FGH', resultados):
            self.sheet_resumo[coluna + str(linha_resumo)] = valor
    
    def coloracao_monitoramento(self):
        
        for linha in range(2, self.sheet_monitoramento.max_row + 1):
            if self.sheet_monitoramento['F' + str(linha)].value == "Sim":

                green_fill = PatternFill(start_color = '66FF66', fill_type = 'solid')

                for coluna in range(1, 7):

                    self.sheet_monitoramento.cell(row = linha, column = coluna).fill = green_fill

            if self.sheet_monitoramento['F' + str(linha)].value == "Não":

                red_fill = PatternFill(start_color = 'FF6666', fill_type = 'solid')

                for coluna in range(1, 7):

                    self.sheet_monitoramento.cell(row = linha, column = coluna).fill = red_fill
     
    
lista_sheets = [campo_r1['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA'], campo_r2['SHEET_DA-PLANILHA']]
lista_regiao = ['R1', 'R2', 'R2', 'R2', 'R2', 'R2']
lista_municipio = ['BRUMADINHO', 'IGARAPÉ', 'BETIM', 'JUATUBA', 'MÁRIO CAMPOS', 'SÃO JOAQUIM DE BICAS']
lista_col_meta = [1,4,7,10,13,16]
lista_linha_resumo = [2,3,4,5,6,7]

for sheet, regiao, municipio, col_meta, linha_resumo in zip(lista_sheets, lista_regiao, lista_municipio, lista_col_meta, lista_linha_resumo):
    
    atualizacao = Atualizacao(sheet, regiao, municipio)
    
    atualizacao.atualizar_meta(col_meta)
    atualizacao.atualizar_controle()
    atualizacao.atualizar_monitoramento()
    atualizacao.coloracao_incompletas()
    atualizacao.situacoes(linha_resumo)
    
coloracao_monitoramento = Atualizacao(None, None, None) 
coloracao_monitoramento.coloracao_monitoramento()

campo_r1.save(caminho_r1)
campo_r2.save(caminho_r2)
campo_r3.save(caminho_r3)
campo_r4.save(caminho_r4)
campo_r5.save(caminho_r5)            

controle.save(caminho_controle)
meta.save(caminho_meta)

os.startfile(caminho_controle)
os.startfile(caminho_meta)

print("Atualização Concluída")
